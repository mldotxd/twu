#!/usr/bin/env python3
"""
测试用例导出工具

将 tc 目录下的用例文件导出为 Excel 格式。
"""

import re
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("运行: uv pip install openpyxl")
    sys.exit(1)


# Excel 表头
HEADERS = [
    "一级分组",
    "二级分组",
    "测试项",
    "优先级",
    "用例标题",
    "前置条件",
    "操作步骤",
    "预期结果",
    "是否反向用例",
    "测试类型",
    "AI生成",
    "备注",
]

# 用例标题正则
TITLE_PATTERN = re.compile(r'^##\s+\[(P[1-5])\](\[反向\])?\s+(.+)$')

# 字段正则
FIELD_PATTERN = re.compile(r'^\[(.+?)\]\s+(.+)$')


def parse_case_file(file_path: Path) -> list[dict]:
    """
    解析用例文件，返回用例列表
    """
    content = file_path.read_text(encoding='utf-8')
    lines = content.split('\n')

    cases = []
    current_case = None

    for line in lines:
        # 匹配用例标题
        title_match = TITLE_PATTERN.match(line)
        if title_match:
            # 保存上一个用例
            if current_case:
                cases.append(current_case)

            priority = title_match.group(1)
            is_negative = title_match.group(2) is not None
            title = title_match.group(3).strip()

            current_case = {
                'priority': priority,
                'is_negative': '是' if is_negative else '否',
                'title': title,
                'test_type': '',
                'precondition': '',
                'steps': '',
                'expected': '',
            }
            continue

        # 匹配字段
        if current_case:
            field_match = FIELD_PATTERN.match(line)
            if field_match:
                field_name = field_match.group(1)
                field_value = field_match.group(2).strip()

                if field_name == '测试类型':
                    current_case['test_type'] = field_value
                elif field_name == '前置条件':
                    current_case['precondition'] = field_value
                elif field_name == '测试步骤':
                    current_case['steps'] = field_value
                elif field_name == '预期结果':
                    current_case['expected'] = field_value

    # 保存最后一个用例
    if current_case:
        cases.append(current_case)

    return cases


def collect_cases(test_case_dir: Path) -> list[dict]:
    """
    收集所有用例
    """
    all_cases = []

    # 遍历 Module 目录
    for module_dir in sorted(test_case_dir.iterdir()):
        if not module_dir.is_dir():
            continue

        module_name = module_dir.name

        # 遍历 Scenario 文件
        for scenario_file in sorted(module_dir.glob('*.md')):
            scenario_name = scenario_file.stem  # 去掉 .md

            # 解析用例
            cases = parse_case_file(scenario_file)

            for case in cases:
                all_cases.append({
                    'module': module_name,
                    'scenario': scenario_name,
                    'test_item': scenario_name,
                    'priority': case['priority'],
                    'title': case['title'],
                    'precondition': case['precondition'],
                    'steps': case['steps'],
                    'expected': case['expected'],
                    'is_negative': case['is_negative'],
                    'test_type': case['test_type'],
                    'ai_generated': '是',
                    'remarks': '',
                })

    return all_cases


def create_excel(cases: list[dict], output_path: Path):
    """
    创建 Excel 文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_alignment = Alignment(vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row, case in enumerate(cases, 2):
        data = [
            case['module'],
            case['scenario'],
            case['test_item'],
            case['priority'],
            case['title'],
            case['precondition'],
            case['steps'],
            case['expected'],
            case['is_negative'],
            case['test_type'],
            case['ai_generated'],
            case['remarks'],
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 设置列宽
    column_widths = [15, 20, 20, 8, 40, 30, 50, 50, 12, 12, 8, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存
    wb.save(output_path)


def main():
    """CLI 入口"""
    # 解析参数
    test_case_dir = None
    output_path = None

    args = sys.argv[1:]
    i = 0
    positional_args = []

    while i < len(args):
        if args[i] in ['-o', '--output']:
            if i + 1 < len(args):
                output_path = Path(args[i + 1])
                i += 2
            else:
                print("错误: --output 需要参数")
                sys.exit(1)
        elif args[i] in ['-h', '--help']:
            print("用法: twu export <tc-dir> [-o output.xlsx]")
            print()
            print("参数:")
            print("  <tc-dir>      tc 目录路径（必填）")
            print()
            print("选项:")
            print("  -o, --output <path>  输出文件路径 (默认: testcase_YYYYMMDD.xlsx)")
            print()
            print("示例:")
            print("  twu export 需求1/tc")
            print("  twu export 需求1/tc -o 需求1_用例.xlsx")
            sys.exit(0)
        elif args[i].startswith('-'):
            print(f"未知参数: {args[i]}")
            sys.exit(1)
        else:
            positional_args.append(args[i])
            i += 1

    # 检查必填参数
    if not positional_args:
        print("错误: 请指定 tc 目录")
        print()
        print("用法: twu export <tc-dir> [-o output.xlsx]")
        print()
        print("示例:")
        print("  twu export 需求1/tc")
        sys.exit(1)

    test_case_dir = Path(positional_args[0])

    # 检查目录
    if not test_case_dir.exists():
        print(f"错误: 目录不存在 {test_case_dir}")
        sys.exit(1)

    # 默认输出文件名
    if output_path is None:
        date_str = datetime.now().strftime('%Y%m%d')
        output_path = Path(f'testcase_{date_str}.xlsx')

    # 收集用例
    print(f"扫描目录: {test_case_dir}")
    cases = collect_cases(test_case_dir)

    if not cases:
        print("警告: 未找到任何用例")
        sys.exit(0)

    # 统计
    modules = set(c['module'] for c in cases)
    scenarios = set((c['module'], c['scenario']) for c in cases)

    print(f"找到 {len(modules)} 个模块，{len(scenarios)} 个场景，{len(cases)} 个用例")

    # 导出
    create_excel(cases, output_path)
    print(f"导出完成: {output_path}")


if __name__ == '__main__':
    main()
